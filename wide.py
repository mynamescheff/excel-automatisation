from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class Wide:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def auto_adjust_column_width(self):
        # Load the Excel workbook
        workbook = load_workbook(self.file_path)
        sheet = workbook[self.sheet_name]

        # Select the whole table
        table_range = sheet.dimensions

        # Auto adjust column width
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Save the modified workbook
        workbook.save(self.file_path)