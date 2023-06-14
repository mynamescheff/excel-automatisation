from openpyxl import load_workbook      
import os
from datetime import date

class ExcelComparator:
    def __init__(self, combined_file, sprawdzacz_file):
        self.combined_file = combined_file
        self.sprawdzacz_file = sprawdzacz_file

    def compare_and_append(self):
        # Load the 'combined2' workbook
        combined_workbook = load_workbook(filename=self.combined_file)
        combined_sheet = combined_workbook['Transposed']

        # Load the 'sprawdzacz' workbook
        sprawdzacz_workbook = load_workbook(filename=self.sprawdzacz_file)
        sprawdzacz_sheet = sprawdzacz_workbook['name_acc']

        # Get the maximum row number in 'combined2' sheet
        max_row = combined_sheet.max_row

        # Create a list to store non-matching values and their corresponding filenames
        non_matching_values = []
        filenames = []

        # Iterate over each row in 'combined2' sheet and check for matches
        for row in range(2, max_row + 1):  # Start from row 2 to skip the header
            case_value = combined_sheet['A' + str(row)].value
            f_value = combined_sheet['F' + str(row)].value
            h_value = combined_sheet['G' + str(row)].value

            # Convert integer values to strings and remove spaces
            f_value = str(f_value).replace(" ", "").replace("-", "") if f_value else ""
            h_value = str(h_value).replace(" ", "").replace("-", "") if h_value else ""

            # Check if any of the values match in 'sprawdzacz' sheet
            found_match = False
            for sprawdzacz_row in sprawdzacz_sheet.iter_rows(values_only=True):
                if f_value == str(sprawdzacz_row[2]) or h_value == str(sprawdzacz_row[2]):
                    found_match = True
                    break

            # If no match found, append the case value, f_value, h_value, and filenames to respective lists
            if not found_match:
                non_matching_values.append((case_value, f_value, h_value))
                filenames.append((self.combined_file, self.sprawdzacz_file))

        # Append the non-matching values and filenames to "mismatch_list.txt"
        file_path = 'C:/IT project/mismatch/mismatch_list.txt'
        with open(file_path, 'w', encoding="utf-8") as file:
            current_date = date.today()
            for value, _ in zip(non_matching_values, filenames):
                modified_value = value[0].replace("\xa0", "").strip()
                file.write(f"{modified_value}: {value[1]}, {value[2]} ({current_date})\n")

        # Print the mismatched values with their corresponding filenames
        print("Mismatched values:")
        for value, _ in zip(non_matching_values, filenames):
            modified_value = value[0].replace("\xa0", "").strip()
            print(f"{modified_value}: {value[1]}, {value[2]} ({current_date})")

        print("Non-matching values appended to mismatch_list.txt successfully.")