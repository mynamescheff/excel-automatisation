import os
from datetime import datetime
from openpyxl import load_workbook

class CaseList:
    def __init__(self, excel_folder, list_folder):
        self.excel_folder = excel_folder
        self.list_folder = list_folder
        self.unique_values = {}

    def process_excel_files(self):
        list_file_path = os.path.join(self.list_folder, "list.txt")

        duplicate_values = set()
        existing_values = set()

        if os.path.isfile(list_file_path):
            # Read the existing values from the file
            existing_values = self.load_existing_list(list_file_path)

        for file_name in os.listdir(self.excel_folder):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(self.excel_folder, file_name)
                try:
                    # Load the Excel file
                    wb = load_workbook(file_path)
                    sheet = wb.active

                    # Extract the value from cell B17
                    value = sheet["B17"].value

                    if value in existing_values:
                        # Check if the value has been previously added
                        duplicate_values.add(value)
                    else:
                        existing_values.add(value)
                        self.unique_values[value] = file_name

                except Exception as e:
                    print(f"Error processing file '{file_path}': {e}")

        if duplicate_values:
            print(f"Alert: Duplicate values found - {duplicate_values}")

        if self.unique_values:
            # Append the unique values to the list file with the current date and time
            today = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(list_file_path, "a") as file:
                file.write(f"\n--- Updated on {today} ---\n")
                for value, file_name in self.unique_values.items():
                    file.write(f"{value} [{file_name}] ({today})\n")

    def load_existing_list(self, list_file_path):
        existing_values = set()

        # Read the existing values from the file
        with open(list_file_path, "r") as file:
            lines = file.readlines()

        # Extract the values from the existing list
        for line in lines:
            line = line.strip()
            if line.startswith("---"):
                # Skip the section headers (e.g., "--- Updated on ... ---")
                continue
            if line:
                parts = line.split(" [")
                value = parts[0]
                existing_values.add(value)

        return existing_values

workbook = Workbook()

worksheet = workbook.active

filename = "C:\\IT project\\test\\combined2.xlsx"
workbook.save(filename)

wb = load_workbook(filename = "C:\\IT project\\test\\combined2.xlsx")

header_list = [
    "uni name",
    "candidate name",
    "case nr",
    "amount",
    "currency",
    "acc number",
    "iban number",
    "swift/bic"
]

ws = wb.active

# Write the header list in column A
for i, header in enumerate(header_list):
    ws[f"A{i+1}"] = header

# Append values to the "combined" excel file starting from column B
for i, excel_file in enumerate(values_excel_files):
    column_letter = get_column_letter(i+2)  # Convert column index to letter
    ws[f"{column_letter}1"] = excel_file  # Write excel file name in the first row of the column
    for j, value in enumerate(values_excel_files[excel_file]):
        # Check if the current value is from cell C34
        if j == len(values_excel_files[excel_file]) - 2:
            # Check if value is not None before replacing spaces
            if value is not None:
                value = value.replace(" ", "")
        ws[f"{column_letter}{j+2}"] = value

wb.save("C:\\IT project\\test\\combined2.xlsx")

filename = "C:\\IT project\\test\\combined2.xlsx"
transposer = ExcelTransposer(filename)
transposer.transpose_cells_to_table()

# Adjust column width
wide = Wide(filename, "Transposed")
wide.auto_adjust_column_width()


excel_folder = "C:\\IT project\\30.05"
list_folder = "C:\\IT project\\case_list"

case_list = CaseList(excel_folder, list_folder)
case_list.process_excel_files()

comparator = ExcelComparator("C:\\IT project\\test\\combined2.xlsx", "C:\\IT project\\sprawdzacz.xlsx")
comparator.compare_and_append()

