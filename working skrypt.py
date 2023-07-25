from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from transposer import ExcelTransposer
from wide import Wide
from case_checker import CaseList
import os
import pandas as pd
from datetime import date
from acc_checker import ExcelComparator
from unidecode import unidecode
from char_map import transform_to_swift_accepted_characters
from charges_checker import check_file_conditions

SOURCE_DIR = "C:\\IT project\\pmt_run"
excel_files = list(Path(SOURCE_DIR).glob("*.xlsx"))

mismatched_cases = []  # List to store cases where the condition isn't met

for excel_file in excel_files:
    wb = load_workbook(filename=excel_file)
    cell_b20_value = wb["Sheet 1"]["B20"].value
    cell_c20_value = wb["Sheet 1"]["C20"].value
    is_condition_met, mismatched_values = check_file_conditions(excel_file.name, cell_b20_value, cell_c20_value)

    if not is_condition_met:
        # Append to mismatched_cases list
        mismatched_cases.append((excel_file.name, mismatched_values))

# Print the cases where the condition isn't met
if mismatched_cases:
    print("Cases with mismatched values in B20 and C20 cells:")
    for case in mismatched_cases:
        print(f"File: {case[0]}, Mismatched Values: {case[1]}")

SOURCE_DIR = "C:\\IT project\\pmt_run"
excel_files = list(Path(SOURCE_DIR).glob("*.xlsx"))

values_excel_files = {}
for excel_file in excel_files:
    wb = load_workbook(filename=excel_file)
    extra_cell_1 = wb["Sheet 1"]["B19"]
    extra_cell_2 = wb["Sheet 1"]["C19"]
    
    if "AQA" in excel_file.name:
        extra_cell_3 = wb["Sheet 1"]["C24"]
        extra_cell_4 = wb["Sheet 1"]["C25"]
        extra_cell_5 = wb["Sheet 1"]["C26"]
    else:
        extra_cell_3 = wb["Sheet 1"]["C33"]
        extra_cell_4 = wb["Sheet 1"]["C34"]
        extra_cell_5 = wb["Sheet 1"]["C35"]
    
    rng_cell_1 = wb["Sheet 1"]["B16"]
    rng_cell_2 = wb["Sheet 1"]["B17"]


    rng_values = [
        rng_cell_1.value,
        rng_cell_2.value
    ]

    # Transform the values to SWIFT-accepted characters using the separate function
    transformed_values = transform_to_swift_accepted_characters(rng_values)  

    extra_cell_1_value = extra_cell_1.value
    extra_cell_2_value = extra_cell_2.value
    extra_cell_3_value = extra_cell_3.value
    extra_cell_4_value = extra_cell_4.value
    extra_cell_5_value = extra_cell_5.value
    transformed_rng_cell_1 = transformed_values[0]
    transformed_rng_cell_2 = transformed_values[1]

    # Add concatenated values to values_excel_files
    values_excel_files[excel_file.name] = [
        transformed_rng_cell_1,
        transformed_rng_cell_2,
        extra_cell_1_value,
        extra_cell_2_value,
        extra_cell_3_value,
        extra_cell_4_value,
        extra_cell_5_value
    ]

workbook = Workbook()

worksheet = workbook.active

filename = "C:\\IT project\\test\\combined2.xlsx"
workbook.save(filename)

wb = load_workbook(filename = "C:\\IT project\\test\\combined2.xlsx")

header_list = [
    "uni name",
    "candidate name",
    "case nr",
    "amount",
    "currency",
    "acc number",
    "iban number",
    "swift/bic"
]

ws = wb.active

# Write the header list in column A
for i, header in enumerate(header_list):
    ws[f"A{i+1}"] = header

# Append values to the "combined" excel file starting from column B
for i, excel_file in enumerate(values_excel_files):
    column_letter = get_column_letter(i+2)  # Convert column index to letter
    ws[f"{column_letter}1"] = excel_file  # Write excel file name in the first row of the column
    for j, value in enumerate(values_excel_files[excel_file]):
        # Check if the current value is from cell C34
        if j == len(values_excel_files[excel_file]) - 2:
            # Check if value is not None and not an integer before replacing spaces
            if value is not None and not isinstance(value, int):
                value = str(value).replace(" ", "")
        ws[f"{column_letter}{j+2}"] = value


wb.save("C:\\IT project\\test\\combined2.xlsx")

filename = "C:\\IT project\\test\\combined2.xlsx"
transposer = ExcelTransposer(filename)
transposer.transpose_cells_to_table()

# Adjust column width
wide = Wide(filename, "Transposed")
wide.auto_adjust_column_width()


excel_folder = "C:\\IT project\\pmt_run"
list_folder = "C:\\IT project\\case_list"

case_list = CaseList(excel_folder, list_folder)
case_list.process_excel_files()

comparator = ExcelComparator("C:\\IT project\\test\\combined2.xlsx", "C:\\IT project\\sprawdzacz.xlsx")
comparator.compare_and_append()