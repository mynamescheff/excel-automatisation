from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelTransposer:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = load_workbook(filename)
        self.sheet = self.workbook.active
    
    def transpose_cells_to_table(self):
        data = []
        
        # Iterate over the rows and columns to collect the data
        for row in self.sheet.iter_rows(values_only=True):
            data.append(row)
        
        # Transpose the data
        transposed_data = list(map(list, zip(*data)))
        
        # Write the transposed data to a new sheet
        transposed_sheet = self.workbook.create_sheet(title="Transposed")
        for row_idx, row_data in enumerate(transposed_data):
            for col_idx, cell_value in enumerate(row_data):
                column_letter = get_column_letter(col_idx + 1)
                transposed_sheet[f"{column_letter}{row_idx + 1}"] = cell_value
        
        # Save the workbook with the transposed data
        self.workbook.save(self.filename)
        print("Transposed data saved successfully.")
